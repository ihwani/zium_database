import openpyxl
import requests

# Pc path
file_path = 'C:/github/zium_database'
# Mac path
# file_path = '/Volumes/Project Drive/zium_database'

excel_file = file_path + '/zium_database.xlsx'
json_file = file_path + '/zium_database_check.json'

wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

sheet = wb.worksheets[0]

url_list = []
for address_row_num in range(2, sheet.max_row):
    url_list.append(sheet.cell(
        row=address_row_num, column=12).value)

check_result_pass = []
check_result_fail = []

for i in range(len(url_list)):
    if requests.get(url_list[i]).status_code == 200:
        result = str(i)+' - pass'
        check_result_pass.append(result)
    else:
        result = str(i)+' - error'
        check_result_fail.append(result)
    print(i)

if len(check_result_fail) == 0:
    print('Sucess')
else:
    print(len(check_result_fail))
    print(check_result_fail)

wb.close
