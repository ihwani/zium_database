import json
import openpyxl

# Pc path
file_path = 'C:/github/zium_database'
# Mac path
# file_path = '/Users/ihwani/Documents/GitHub/zium_database'

excel_file = file_path + '/zium_database.xlsx'
json_file = file_path + '/zium_database.json'

wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

sheet = wb.worksheets[0]

key_list = []
for col_num in range(1, sheet.max_column + 1):
    key_list.append(sheet.cell(row=1, column=col_num).value)

data_dict = []
for row_num in range(2, sheet.max_row + 1):
    tmp_dict = {}
    tag = ""
    boss = ""
    for col_num in range(1, sheet.max_column + 1):
        val = sheet.cell(row=row_num, column=col_num).value
        if col_num == 6:
            boss = sheet.cell(row=row_num, column=col_num).value
        elif col_num < sheet.max_column:
            tmp_dict[key_list[col_num - 1]] = val
        else:
            tag = sheet.cell(row=row_num, column=col_num).value
    boss = boss.split(',')
    boss = [line.strip() for line in boss]
    tag = tag.split(',')
    tag = [line.strip() for line in tag]
    tmp_dict[key_list[5]] = boss
    tmp_dict[key_list[-1]] = tag
    data_dict.append(tmp_dict)
    print(len(data_dict))
print(str(len(data_dict)) + "개의 DB가 등록되었습니다.")

wb.close()

with open(json_file, 'w', encoding='utf-8') as fp:
    json.dump(data_dict, fp, indent=4, ensure_ascii=False)

excel_file = file_path + '/zium_database.xlsx'
json_file = file_path + '/zium_office_address.json'

wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

sheet = wb.worksheets[1]

address_key_list = []
for address_row_num in range(2, sheet.max_row + 1):
    address_key_list.append(sheet.cell(row=address_row_num, column=1).value)

address_value_list = []
for address_row_num in range(2, sheet.max_row + 1):
    address_value_list.append(sheet.cell(
        row=address_row_num, column=sheet.max_row).value)

address_data_dict = {}
for i in range(len(address_key_list)):
    address_data_dict[address_key_list[i]] = address_value_list[i]

print(str(len(address_data_dict)) + "개의 주소가 등록되었습니다..")

wb.close()

with open(json_file, 'w', encoding='utf-8') as fp:
    json.dump(address_data_dict, fp, indent=4, ensure_ascii=False)
